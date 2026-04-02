# -*- coding: utf-8 -*-
"""
플레이오토 주문 데이터 자동 수집기 (EXE용)
시스템에 설치된 Chrome 브라우저 사용
"""

import sys
import os
import json
import time
import logging
from datetime import datetime, timezone, timedelta
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import requests
import base64

# KST 타임존
KST = timezone(timedelta(hours=9))

def now_kst():
    return datetime.now(KST)

# 실행 파일 기준 경로
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

# 설정 파일 경로
CONFIG_FILE = BASE_DIR / "config.json"
LOG_DIR = BASE_DIR / "logs"
DOWNLOAD_DIR = BASE_DIR / "downloads"

LOG_DIR.mkdir(exist_ok=True)
DOWNLOAD_DIR.mkdir(exist_ok=True)

# 로깅 설정
log_file = LOG_DIR / f"collector_{now_kst().strftime('%Y%m%d')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def load_config():
    """설정 파일 로드"""
    if not CONFIG_FILE.exists():
        # 기본 설정 파일 생성
        default_config = {
            "playauto_id": "",
            "playauto_pw": "",
            "github_token": "",
            "github_repo": "692meatlab/playauto-analyzer",
            "headless": True
        }
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, indent=2, ensure_ascii=False)
        logger.error(f"설정 파일이 생성되었습니다: {CONFIG_FILE}")
        logger.error("config.json 파일을 열어 로그인 정보를 입력하세요.")
        return None

    with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


class PlayautoCollector:
    """플레이오토 데이터 수집기"""

    def __init__(self, config):
        self.config = config
        self.browser = None
        self.page = None
        self.downloaded_file = None

    def run(self):
        """전체 수집 프로세스 실행"""
        logger.info("=" * 50)
        logger.info("플레이오토 데이터 수집 시작")
        logger.info(f"시작 시간: {now_kst().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 50)

        try:
            with sync_playwright() as p:
                # 시스템 Chrome 사용
                self.browser = p.chromium.launch(
                    headless=self.config.get('headless', True),
                    channel='chrome',  # 시스템 Chrome 사용
                    args=['--lang=ko-KR']
                )
                context = self.browser.new_context(
                    accept_downloads=True,
                    locale='ko-KR'
                )
                self.page = context.new_page()
                self.page.set_default_timeout(60000)

                # 1. 로그인
                if not self.login():
                    raise Exception("로그인 실패")

                # 2. 수집 버튼 클릭
                self.click_collect()

                # 3. 주문 페이지로 이동
                if not self.go_to_orders():
                    raise Exception("주문 페이지 이동 실패")

                # 4. 엑셀 다운로드
                if not self.download_excel():
                    raise Exception("엑셀 다운로드 실패")

                # 5. GitHub 업로드
                if not self.upload_to_github():
                    raise Exception("GitHub 업로드 실패")

                logger.info("=" * 50)
                logger.info("데이터 수집 완료!")
                logger.info("=" * 50)
                return True

        except Exception as e:
            logger.error(f"수집 실패: {e}")
            return False

        finally:
            if self.browser:
                try:
                    self.browser.close()
                except Exception:
                    pass  # Python 3.14 event loop 종료 시 발생하는 비치명적 오류 무시

    def login(self):
        """플레이오토 로그인"""
        logger.info("로그인 시도...")

        try:
            self.page.goto("https://app.playauto.io/login.html")
            self.page.wait_for_load_state('networkidle')
            time.sleep(2)

            # 아이디 입력
            id_input = self.page.wait_for_selector(
                'input[type="email"], input[type="text"], input[name="email"]',
                timeout=10000
            )
            if id_input:
                id_input.fill(self.config['playauto_id'])
                logger.info("아이디 입력 완료")

            # 비밀번호 입력
            pw_input = self.page.wait_for_selector('input[type="password"]', timeout=5000)
            if pw_input:
                pw_input.fill(self.config['playauto_pw'])
                logger.info("비밀번호 입력 완료")

            # 로그인 버튼 클릭
            login_btn = self.page.wait_for_selector(
                'button[type="submit"], button:has-text("Sign In"), button:has-text("로그인")',
                timeout=5000
            )
            if login_btn:
                login_btn.click()

            time.sleep(5)
            self.page.wait_for_load_state('networkidle')

            # 로그인 성공 판단: 대시보드 요소 확인
            try:
                # 왼쪽 메뉴나 대시보드 요소가 있으면 성공
                dashboard = self.page.wait_for_selector(
                    'text=주문, text=결제관리, text=판매자센터, .sidebar, nav',
                    timeout=5000
                )
                if dashboard:
                    logger.info("로그인 성공!")
                    self.page.screenshot(path=str(LOG_DIR / "login_success.png"))
                    return True
            except:
                pass

            # URL 체크 (fallback)
            if 'login' not in self.page.url.lower():
                logger.info("로그인 성공!")
                return True

            logger.error("로그인 실패")
            self.page.screenshot(path=str(LOG_DIR / "login_failed.png"))
            return False

        except Exception as e:
            logger.error(f"로그인 오류: {e}")
            self.page.screenshot(path=str(LOG_DIR / "login_error.png"))
            return False

    def click_collect(self):
        """수집 버튼 클릭"""
        logger.info("수집 버튼 찾는 중...")
        try:
            btn = self.page.wait_for_selector(
                'button:has-text("수집"), button:has-text("주문수집")',
                timeout=5000
            )
            if btn and btn.is_visible():
                btn.click()
                logger.info("수집 버튼 클릭")
                time.sleep(5)
                return True
        except:
            logger.info("수집 버튼 없음 (이미 수집됨)")
        return False

    def close_popups(self):
        """팝업 닫기.

        플레이오토에는 두 종류의 팝업이 겹쳐서 뜸:
        1. "사업자정보 업데이트 안내" 모달 → "변경사항 없음" 클릭
        2. 광고/공지 팝업 → "다음" 반복 후 "닫기" 활성화되면 클릭

        처리 순서가 중요: 앞 팝업(사업자정보)을 먼저 닫아야
        뒤 팝업(광고)의 버튼을 클릭할 수 있음.
        """
        logger.info("팝업 확인 및 닫기...")
        time.sleep(2)

        # 1단계: 사업자정보 업데이트 안내 모달 닫기
        for selector in [
            "button:has-text('변경사항 없음')",
            "button:has-text('2주간 표시 안함')",
        ]:
            try:
                btn = self.page.locator(selector).first
                if btn.is_visible(timeout=1000) and btn.is_enabled():
                    btn.click()
                    logger.info(f"사업자정보 모달 닫기: {selector}")
                    time.sleep(1)
                    break
            except Exception:
                pass

        # 2단계: 비밀번호 변경 안내 모달 닫기
        for selector in [
            "button:has-text('다음에 변경하기')",
            "button:has-text('2주간 표시 안함')",
        ]:
            try:
                btn = self.page.locator(selector).first
                if btn.is_visible(timeout=1000) and btn.is_enabled():
                    btn.click()
                    logger.info(f"비밀번호 모달 닫기: {selector}")
                    time.sleep(1)
                    break
            except Exception:
                pass

        # 2단계: 광고/공지 팝업 처리 (다음 반복 → 닫기)
        iterations = 0
        max_iterations = 30

        while iterations < max_iterations:
            iterations += 1
            action_taken = False

            contexts = [self.page] + [
                f for f in self.page.frames if f != self.page.main_frame
            ]

            for ctx in contexts:
                # 우선순위 1: "다음" 버튼 클릭 (닫기 활성화용)
                for selector in [
                    "button:has-text('다음')",
                    "a:has-text('다음')",
                ]:
                    try:
                        btn = ctx.locator(selector).first
                        if btn.is_visible(timeout=300) and btn.is_enabled():
                            btn.click()
                            logger.info(f"광고 팝업 '다음' 클릭 ({iterations}번째)")
                            time.sleep(0.6)
                            action_taken = True
                            break
                    except Exception:
                        pass
                if action_taken:
                    break

                # 우선순위 2: 활성화된 "닫기" 클릭
                for selector in [
                    "button:has-text('닫기')",
                    "a:has-text('닫기')",
                    "button:has-text('오늘하루 보지않기')",
                    "a:has-text('오늘하루 보지않기')",
                    ".popup-close",
                    ".modal-close",
                ]:
                    try:
                        btn = ctx.locator(selector).first
                        if btn.is_visible(timeout=300) and btn.is_enabled():
                            btn.click()
                            logger.info(f"광고 팝업 닫기 ({iterations}번째)")
                            time.sleep(0.6)
                            action_taken = True
                            break
                    except Exception:
                        pass
                if action_taken:
                    break

            if not action_taken:
                logger.info(f"팝업 처리 완료 (총 {iterations - 1}회)")
                break

        if iterations >= max_iterations:
            logger.warning("팝업 처리 최대 반복 도달")
            self.page.screenshot(path=str(LOG_DIR / "popup_timeout.png"))

        # 3단계: modal-company-dim 오버레이 강제 제거 (클릭 차단 해제)
        try:
            overlay = self.page.locator(".modal-dim, .modal-company-dim").first
            if overlay.is_visible(timeout=1000):
                logger.info("modal-dim 오버레이 감지 — Escape 키 시도")
                self.page.keyboard.press("Escape")
                time.sleep(1)
        except Exception:
            pass

        try:
            still_blocking = self.page.locator(".modal-dim, .modal-company-dim").first
            if still_blocking.is_visible(timeout=1000):
                logger.info("Escape 후에도 오버레이 잔존 — JS로 강제 제거")
                self.page.evaluate("""() => {
                    document.querySelectorAll('.modal-dim, .modal-company-dim').forEach(el => el.remove());
                }""")
                time.sleep(0.5)
        except Exception:
            pass

    def go_to_orders(self):
        """주문 > 전체 주문 조회 페이지로 이동"""
        logger.info("주문 페이지 이동...")

        # 팝업 먼저 닫기
        self.close_popups()
        time.sleep(1)

        try:
            self.page.screenshot(path=str(LOG_DIR / "step1_before_menu.png"))

            # 1. 좌측 "주문" 메뉴 클릭 (팝업 재등장 시 닫고 재시도)
            logger.info("1단계: 주문 메뉴 클릭")
            for attempt in range(3):
                try:
                    self.page.locator("#navi-order").click(timeout=10000)
                    break
                except Exception:
                    logger.info(f"메뉴 클릭 실패, 팝업 재처리 후 재시도 ({attempt+1}/3)")
                    self.close_popups()
            time.sleep(1.5)
            self.page.screenshot(path=str(LOG_DIR / "step2_after_order.png"))

            # 2. "전체 주문 조회" 클릭
            logger.info("2단계: 전체 주문 조회 클릭")
            for attempt in range(3):
                try:
                    self.page.get_by_text("전체 주문 조회", exact=True).click(timeout=10000)
                    break
                except Exception:
                    logger.info(f"메뉴 항목 클릭 실패, 팝업 재처리 후 재시도 ({attempt+1}/3)")
                    self.close_popups()
                    self.page.locator("#navi-order").click(timeout=5000)
                    time.sleep(1)
            self.page.wait_for_load_state('networkidle')
            time.sleep(2)

            # 이동 후 팝업 재확인
            self.close_popups()

            self.page.screenshot(path=str(LOG_DIR / "step3_order_list.png"))
            logger.info(f"현재 URL: {self.page.url}")
            return True

        except Exception as e:
            logger.error(f"주문 페이지 이동 실패: {e}")
            self.page.screenshot(path=str(LOG_DIR / "order_nav_failed.png"))
            return False

    def download_excel(self):
        """엑셀 다운로드: 어제날짜 설정 > 검색 > 엑셀 > 통합 엑셀 다운 > 다운로드"""
        logger.info("엑셀 다운로드 시작...")

        try:
            # 3. "어제" 버튼 클릭 (전날 데이터)
            logger.info("3단계: 날짜 '어제' 설정")
            self.page.get_by_text("어제", exact=True).first.click()
            time.sleep(1)
            logger.info("날짜 '어제' 설정 완료")

            # 4. 검색 버튼 클릭
            logger.info("4단계: 검색 버튼 클릭")
            self.page.get_by_text("검색", exact=True).first.click()
            time.sleep(3)
            self.page.wait_for_load_state('networkidle')
            logger.info("검색 완료")
            self.page.screenshot(path=str(LOG_DIR / "step4_after_search.png"))

            # 5. 엑셀 드롭다운 클릭
            logger.info("5단계: 엑셀 버튼 클릭")
            self.page.get_by_text("엑셀", exact=True).first.click()
            time.sleep(1)
            self.page.screenshot(path=str(LOG_DIR / "step5_excel_menu.png"))

            # 6. "통합 엑셀 다운" 클릭
            logger.info("6단계: 통합 엑셀 다운 클릭")
            self.page.get_by_text("통합 엑셀 다운", exact=True).click()
            time.sleep(2)
            self.page.screenshot(path=str(LOG_DIR / "step6_download_modal.png"))
            logger.info("통합 엑셀 다운 클릭 완료")

            # 6.5. 양식 행 선택 - Playwright 네이티브 클릭 (Angular 이벤트 정상 발화)
            logger.info("6.5단계: 양식 행 선택")
            self.page.screenshot(path=str(LOG_DIR / "step65_before_row.png"))

            row_clicked = False
            try:
                # Playwright locator로 ui-grid-row 클릭 (JS .click()보다 이벤트 정확)
                row_locator = self.page.locator("div.ui-grid-row").filter(has_text="엑셀다운").first
                row_locator.wait_for(state="visible", timeout=5000)
                row_locator.click()
                logger.info("ui-grid-row 클릭 완료 (Playwright locator)")
                row_clicked = True
                time.sleep(0.8)
            except Exception as e:
                logger.warning(f"locator 클릭 실패: {e}, JS fallback 시도")
                # fallback: classes.includes()로 정확히 'ui-grid-row' 매칭
                result = self.page.evaluate("""() => {
                    const cells = document.querySelectorAll('.ui-grid-cell-contents');
                    for (const cell of cells) {
                        if (cell.textContent.trim() === '엑셀다운') {
                            let node = cell.parentElement;
                            while (node) {
                                const classes = (node.className || '').toString().trim().split(' ');
                                if (classes.includes('ui-grid-row')) {
                                    node.click();
                                    return 'js-clicked:' + node.className.toString().slice(0, 50);
                                }
                                node = node.parentElement;
                            }
                            cell.click();
                            return 'cell-fallback';
                        }
                    }
                    return 'not-found';
                }""")
                logger.info(f"JS fallback 결과: {result}")
                row_clicked = True
                time.sleep(0.8)

            self.page.screenshot(path=str(LOG_DIR / "step65_after_row.png"))

            # 7. 다운로드 버튼 클릭
            # Angular download() → window.open(url) 호출
            # Chrome 팝업 차단기가 막을 수 있으므로 JS로 window.open을 가로채
            # URL을 캡처한 뒤 requests로 직접 다운로드
            logger.info("7단계: 다운로드 버튼 클릭")
            download_btn = self.page.locator("button.btn-primary:has-text('다운로드')")
            download_btn.scroll_into_view_if_needed()
            time.sleep(0.3)
            self.page.screenshot(path=str(LOG_DIR / "step7_before_download.png"))

            # window.open 가로채기 (팝업 차단 우회)
            # location.assign / location.href 변경도 함께 감지
            self.page.evaluate("""() => {
                window._playautoDownloadUrl = null;
                window._playautoConsoleLog = [];
                const _origOpen = window.open;
                window.open = function(url) {
                    window._playautoDownloadUrl = url || '__no_url__';
                    console.log('[PA] window.open intercepted:', url);
                    return null;
                };
                // location.assign 도 감지
                const _origAssign = window.location.assign.bind(window.location);
                try {
                    Object.defineProperty(window.location, 'assign', {
                        value: function(url) {
                            if (url && (url.includes('.xlsx') || url.includes('download') || url.includes('excel'))) {
                                window._playautoDownloadUrl = url;
                                console.log('[PA] location.assign intercepted:', url);
                            } else {
                                _origAssign(url);
                            }
                        },
                        writable: true
                    });
                } catch(e) {}
            }""")

            # 콘솔 메시지 수집 (디버깅용)
            console_msgs = []
            self.page.on("console", lambda msg: console_msgs.append(f"[{msg.type}] {msg.text}"))

            download_btn.click()
            logger.info("다운로드 버튼 클릭 완료")

            # "양식을 선택해 주십시오." 알림창 감지 및 처리
            time.sleep(0.8)
            try:
                alert_ok = self.page.locator("div.modal-content button:has-text('확인'), .sweet-alert button:has-text('확인'), button:has-text('확인')").first
                if alert_ok.is_visible(timeout=1000):
                    logger.warning("'양식을 선택해 주십시오.' 알림 감지 - 행 재선택 시도")
                    alert_ok.click()
                    time.sleep(0.5)

                    # 행 재선택 후 다운로드 재시도
                    try:
                        row_locator2 = self.page.locator("div.ui-grid-row").filter(has_text="엑셀다운").first
                        row_locator2.click()
                        logger.info("행 재선택 완료")
                        time.sleep(0.8)
                    except Exception as e2:
                        logger.error(f"행 재선택 실패: {e2}")
                        self.page.screenshot(path=str(LOG_DIR / "step7_reselect_failed.png"))
                        return False

                    # 다운로드 재시도
                    download_btn2 = self.page.locator("button.btn-primary:has-text('다운로드')")
                    download_btn2.click()
                    logger.info("다운로드 버튼 재클릭")
                    time.sleep(0.5)
            except Exception:
                pass  # 알림창 없으면 정상

            # Angular는 비동기로 서버 요청 후 window.open 호출할 수 있으므로 최대 30초 대기
            dl_url = None
            for i in range(60):
                time.sleep(0.5)
                dl_url = self.page.evaluate("() => window._playautoDownloadUrl")
                if dl_url:
                    break
                if i % 10 == 9:
                    logger.info(f"window.open 대기 중... ({(i+1)//2}초)")
                    self.page.screenshot(path=str(LOG_DIR / f"step7_wait_{(i+1)//2}s.png"))

            # 콘솔 로그 기록 (Angular 에러 진단)
            pa_logs = [m for m in console_msgs if '[PA]' in m or 'error' in m.lower()]
            if pa_logs:
                logger.info(f"콘솔 로그: {pa_logs[-5:]}")

            if not dl_url:
                logger.error("window.open URL을 감지하지 못함 - 스크린샷 저장")
                logger.error(f"전체 콘솔 로그 (마지막 10개): {console_msgs[-10:]}")
                self.page.screenshot(path=str(LOG_DIR / "step7_no_url.png"))
                return False

            logger.info(f"다운로드 URL 감지: {dl_url}")

            # 브라우저 세션 쿠키로 인증된 요청
            raw_cookies = self.page.context.cookies()
            session = requests.Session()
            for c in raw_cookies:
                session.cookies.set(c['name'], c['value'], domain=c.get('domain', ''))

            resp = session.get(dl_url, timeout=120, stream=True)
            if resp.status_code != 200:
                logger.error(f"다운로드 요청 실패: {resp.status_code}")
                return False

            timestamp = now_kst().strftime('%Y%m%d_%H%M%S')
            save_path = DOWNLOAD_DIR / f"playauto_{timestamp}.xlsx"
            with open(save_path, 'wb') as f:
                for chunk in resp.iter_content(chunk_size=8192):
                    f.write(chunk)

            logger.info(f"전체 다운로드 완료: {save_path} ({save_path.stat().st_size:,} bytes)")

            # 어제 날짜 기준으로 필터링 (주문수집일 컬럼)
            filtered_path = self._filter_yesterday(save_path)
            if filtered_path:
                self.downloaded_file = filtered_path
            else:
                # 필터 실패 시 원본 사용
                self.downloaded_file = save_path
            return True

        except PlaywrightTimeout:
            logger.error("다운로드 시간 초과")
            self.page.screenshot(path=str(LOG_DIR / "download_timeout.png"))
            return False
        except Exception as e:
            logger.error(f"다운로드 오류: {e}")
            self.page.screenshot(path=str(LOG_DIR / "download_error.png"))
            return False

    def _filter_yesterday(self, src_path: Path) -> Path | None:
        """전체 엑셀에서 어제 주문수집일 행만 추출해 별도 파일로 저장"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            yesterday = (now_kst() - timedelta(days=1)).date()
            logger.info(f"어제 날짜 필터: {yesterday}")

            wb = load_workbook(src_path)
            ws = wb.active

            # 헤더에서 '주문수집일' 컬럼 위치 찾기
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            col_idx = None
            for i, h in enumerate(headers):
                if h and '주문수집일' in str(h):
                    col_idx = i + 1
                    break

            if col_idx is None:
                logger.warning("'주문수집일' 컬럼을 찾지 못함 - 필터 건너뜀")
                return None

            logger.info(f"'주문수집일' 컬럼: {col_idx}번째")

            # 어제 날짜 행만 추출
            keep_rows = [list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]  # 헤더
            filtered = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col_idx - 1]
                if val is None:
                    continue
                # datetime 또는 문자열 모두 처리
                if hasattr(val, 'date'):
                    row_date = val.date()
                else:
                    try:
                        row_date = datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
                    except Exception:
                        continue
                if row_date == yesterday:
                    keep_rows.append(row)
                    filtered += 1

            logger.info(f"필터 결과: 전체 {ws.max_row - 1}건 → 어제({yesterday}) {filtered}건")

            # 새 워크북에 저장
            from openpyxl import Workbook
            new_wb = Workbook()
            new_ws = new_wb.active
            for r in keep_rows:
                new_ws.append(list(r))

            yesterday_str = yesterday.strftime('%Y%m%d')
            filtered_path = src_path.parent / f"playauto_{yesterday_str}_filtered.xlsx"
            new_wb.save(filtered_path)
            logger.info(f"필터 파일 저장: {filtered_path} ({filtered_path.stat().st_size:,} bytes)")
            return filtered_path

        except Exception as e:
            logger.error(f"필터링 오류: {e}")
            return None

    def upload_to_github(self):
        """GitHub 업로드 - 업로드별 개별 파일로 저장"""
        logger.info("GitHub 업로드...")

        if not self.downloaded_file or not self.downloaded_file.exists():
            logger.error("업로드 파일 없음")
            return False

        try:
            from openpyxl import load_workbook
            import csv
            from io import StringIO

            wb = load_workbook(self.downloaded_file)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                logger.error("새 데이터 없음")
                return False

            col_headers = all_rows[0]
            new_data = all_rows[1:]

            api_headers = {
                "Authorization": f"token {self.config['github_token']}",
                "Accept": "application/vnd.github.v3+json"
            }
            repo = self.config['github_repo']

            # 업로드 ID 생성
            upload_id = now_kst().strftime('%Y%m%d_%H%M%S') + '_auto'

            # CSV 생성
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(col_headers)
            for row in new_data:
                writer.writerow([str(v) if v is not None else '' for v in row])
            csv_content = output.getvalue()

            # 개별 파일로 업로드
            upload_url = f"https://api.github.com/repos/{repo}/contents/data/uploads/{upload_id}.csv"
            content_b64 = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
            resp = requests.put(upload_url, headers=api_headers, json={
                "message": f"자동 수집: {now_kst().strftime('%Y-%m-%d %H:%M')}",
                "content": content_b64,
                "branch": "main"
            })
            if resp.status_code not in [200, 201]:
                logger.error(f"데이터 파일 업로드 실패: {resp.status_code}")
                return False

            logger.info(f"데이터 파일 업로드 성공: {upload_id}.csv ({len(new_data)}건)")

            # 날짜 범위 계산 (결제완료일 컬럼 기준)
            col_list = [str(h) if h is not None else '' for h in col_headers]
            date_col_idx = next((i for i, h in enumerate(col_list) if '결제완료일' in h), None)
            dates = []
            if date_col_idx is not None:
                for row in new_data:
                    val = row[date_col_idx] if len(row) > date_col_idx else None
                    if val:
                        try:
                            d = str(val)[:10]
                            dates.append(d)
                        except Exception:
                            pass
            start_date = min(dates) if dates else now_kst().strftime('%Y-%m-%d')
            end_date = max(dates) if dates else start_date

            # metadata.json 업데이트
            meta_url = f"https://api.github.com/repos/{repo}/contents/data/metadata.json"
            meta_resp = requests.get(meta_url, headers=api_headers)
            meta_sha = None
            metadata = {}
            if meta_resp.status_code == 200:
                meta_sha = meta_resp.json().get('sha')
                metadata = json.loads(base64.b64decode(meta_resp.json()['content']).decode('utf-8'))

            metadata[upload_id] = {
                'file_name': 'auto-collected',
                'period': f"{start_date} ~ {end_date}",
                'start_date': start_date,
                'end_date': end_date,
                'row_count': len(new_data),
                'order_count': len(new_data),
                'total_revenue': 0,
                'cancel_count': 0,
                'loaded_at': now_kst().strftime('%Y-%m-%d %H:%M:%S')
            }

            meta_content = base64.b64encode(
                json.dumps(metadata, ensure_ascii=False, indent=2).encode('utf-8')
            ).decode('utf-8')
            meta_payload = {
                "message": f"메타데이터 업데이트: {upload_id}",
                "content": meta_content,
                "branch": "main"
            }
            if meta_sha:
                meta_payload["sha"] = meta_sha
            requests.put(meta_url, headers=api_headers, json=meta_payload)

            logger.info("GitHub 업로드 완료!")
            return True

        except ImportError:
            logger.error("openpyxl 패키지 필요")
            return False
        except Exception as e:
            logger.error(f"업로드 오류: {e}")
            return False


def main():
    print("=" * 50)
    print("플레이오토 자동 수집기")
    print("=" * 50)
    print()

    config = load_config()
    if not config:
        input("Enter 키를 눌러 종료...")
        sys.exit(1)

    if not config.get('playauto_id') or not config.get('playauto_pw'):
        logger.error("config.json에 playauto_id, playauto_pw를 입력하세요")
        input("Enter 키를 눌러 종료...")
        sys.exit(1)

    if not config.get('github_token'):
        logger.error("config.json에 github_token을 입력하세요")
        input("Enter 키를 눌러 종료...")
        sys.exit(1)

    collector = PlayautoCollector(config)
    success = collector.run()

    if not success:
        input("Enter 키를 눌러 종료...")

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
